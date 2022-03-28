# Functions for filling in Excel Template files used for Alpha ATE LabVIEW
# softwares
# Date: March 28th, 2022
# Written By: Tristan Lee
from openpyxl import load_workbook
import win32com.client


# Fills in ATE excel template of file type .xlsx
# requires: results of test_names are at corresponding indeces in test_results
def write_results(template_path, results_path, test_names, test_results):

    # loading template file
    report = load_workbook(template_path)
    report.template = False  # this is needed to properly save an xlsx file
    sheet = report["A5 form"]  # selecting the sheet within the workbook

    # iterates through spreadsheet cells
    for r in range(2, sheet.max_row+1):
        for c in range(3, sheet.max_column+1):
            val = sheet.cell(r, c).value  # value of cell
            if val != None:  # check to ensure the cell is not empty
                # test_names indeces align with test_results'
                for idx, name in enumerate(test_names):
                    if name == str(val):
                        try:
                            # conversion necassary for formulae
                            sheet.cell(r, c).value = float(test_results[idx])
                        except:
                            sheet.cell(r, c).value = test_results[idx]

    report.save(results_path)  # saves filled in workbook file


# Generates a PDF from a given results_path xlsx file
def generate_pdf(results_path, pdf_path):
    # opens excel
    excel = win32com.client.Dispatch("Excel.Application")

    # opens excel workbook file
    final_report = excel.Workbooks.Open(results_path)
    sheet_indeces = [1]

    # saves excel file as pdf and closes
    final_report.WorkSheets(sheet_indeces).Select()
    final_report.ActiveSheet.ExportAsFixedFormat(0, pdf_path)
    final_report.Close()
